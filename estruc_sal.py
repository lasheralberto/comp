# -*- coding: utf-8 -*-
"""estruc_sal.ipynb

Automatically generated by Colaboratory.

Original file is located at
    https://colab.research.google.com/drive/17oMXxJ77UH9f8kh8Pd3SFqQoaf_TaSMB
"""

import pandas as pd 
import plotly.express as px
import seaborn as sns
import matplotlib.pyplot as plt
from plotly.subplots import make_subplots
from openpyxl import load_workbook

!gdown --id 163MC5wZ2ohH3a95GrOz1m03CLTHUK8Vh
#https://drive.google.com/file/d/163MC5wZ2ohH3a95GrOz1m03CLTHUK8Vh/view?usp=sharing

df=pd.read_excel('PREPARATIVO POWER BI.xlsx')
df=df.fillna(value=0)
cols=['Cód. trabajador','Nombre trabajador','Categoría','Tipo contrato','DEPARTAMENTO','temporalidad','Itx','Antigüedad','Sexo']
dfnum=df.drop(columns=cols)
numcols=dfnum.columns
df[numcols]=df[numcols].apply(pd.to_numeric)

#ESPECIFICAMOS EL NOMBRE DEL ARCHIVO DONDE GUARDAMOS LOS DATOS
archivo_nombre='estruc.xlsx'

#si le indicas dni te elimina las columnas y los complementos que son 0. Si no lo indicas te aparecerá la categoría entera pero sólo las columnas que tengan todo 0 se eliminarán
def complementos_find(categoria, dni):
  df_comp=pd.pivot_table(df, index=['Categoría', 'Cód. trabajador'])
  df_comp=df_comp.loc[categoria]
  df_comp=df_comp.drop(columns=['Diferencia.1','Horas semanales','SALARIO BASE','CALCULADO','Ant. Años','Diferencia','Tot. Devengado','Total salario complementos','Total salario convenio ','Complementos convenio','Complementos no convenio'])
  df_comp=df_comp.loc[:, (df_comp != 0).any(axis=0)]
  

  if dni!='':
    df_comp=df_comp.loc[dni].to_frame('Complementos')
  
    for i in df_comp['Complementos']:
      if i==0:
        df_comp=df_comp.loc[(df_comp!=0).any(axis=1)]
  
  return df_comp

complementos_find('ANALISTA','')

#agrupación del total devengado por intervalos entre Hombres y Mujeres
df_sex=df.set_index('Sexo')
df_men=df_sex.loc['Hombre']
df_women=df_sex.loc['Mujer']

dfmen_intervals=df_men.groupby(pd.cut(df_men['Tot. Devengado'],[0,1200,2400,3600,4800,6000])).count()
dfmen_intervals=dfmen_intervals['Tot. Devengado']
dfmen_intervals=dfmen_intervals.to_frame('Hombres, Tot. Devengado')
dfmen_intervals['fi%']=[round(i/dfmen_intervals.sum(),2) for i in dfmen_intervals['Hombres, Tot. Devengado'] ]
dfmen_intervals['fi%']=dfmen_intervals['fi%'].astype('float')

dfwomen_intervals=df_women.groupby(pd.cut(df_women['Tot. Devengado'],[0,1200,2400,3600,4800,6000])).count()
dfwomen_intervals=dfwomen_intervals['Tot. Devengado']
dfwomen_intervals=dfwomen_intervals.to_frame('Mujeres, Tot. Devengado')
dfwomen_intervals['fi%']=[round(i/dfwomen_intervals.sum(),2) for i in dfwomen_intervals['Mujeres, Tot. Devengado'] ]
dfwomen_intervals['fi%']=dfwomen_intervals['fi%'].astype('float')

df_intervals=pd.merge(dfmen_intervals,dfwomen_intervals,on=dfmen_intervals.index)
df_intervals.rename(columns={'key_0':'Intervalos','fi%_x':'%Hombres','fi%_y':'%Mujeres'})

writer= pd.ExcelWriter(archivo_nombre, engine='openpyxl')
df_intervals.to_excel(writer, sheet_name='Intervalos', index=False)
writer.save()
writer.close()

#LISTADO EN EXCEL
def generar_diferencias_complementosNoConvenio(df):
#tabla para hombres

  df_men_bycat_compl_convenio=pd.pivot_table(df, index='Categoría', aggfunc='mean')[['Total salario convenio ','Complementos convenio', 'Complementos no convenio']]
  df_men_bycat_compl_convenio=df_men_bycat_compl_convenio.reset_index()
#juntamos los promedios por categoría

  df_men1=df.merge(df_men_bycat_compl_convenio, on='Categoría')
  df_men1=df_men1.rename(columns={'Total salario convenio _x':'Total salario convenio','Complementos convenio_x':'Complementos convenio','Complementos no convenio_x':'Complementos no convenio','Total salario convenio _y':'salario convenio media categoría','Complementos convenio_y':'Complementos convenio media categoría','Complementos no convenio_y':'Complementos no convenio media categoría'})
  df_men1=df_men1[['Cód. trabajador','Nombre trabajador','Categoría','DEPARTAMENTO','Ant. Años','Tot. Devengado','Total salario complementos','Total salario convenio','Complementos convenio','Complementos no convenio', 'Complementos convenio media categoría','Complementos no convenio media categoría']]
#cálculo de la diferencia entre salario convenio y el promedio (por CATEGORIA) de complementos no convenio 

  df_men1['Diferencia_SC_CNCProm']=df_men1['Complementos no convenio']-df_men1['Complementos no convenio media categoría']
#le ponemos el filtro para saber quién tiene diferencias positivas. Si no filtramos nos saldrán aquellos que no tengan complementos no covenio. Comparar la brecha de los complementos no convenio con aquellos que no tienen complementos no convenio no tiene sentido.
  filtrados_men=df_men1[df_men1['Complementos no convenio']>df_men1['Complementos no convenio media categoría']] 
  #filtrados_men=df_men1 #sin filtrar
  with pd.ExcelWriter(archivo_nombre, engine='openpyxl') as writer1:
    writer1.book = load_workbook(archivo_nombre)
    filtrados_men.to_excel(writer1, sheet_name="Diferencias comp. no convenio", index=False)
    writer1.save()
  filtrados_men.to_excel('df.xlsx',engine='openpyxl')
  return filtrados_men.head()
generar_diferencias_complementosNoConvenio(df)

#COMPLEMENTOS CONVENIO. TOTAL
df2=df.set_index('Sexo')
df_h=df2.loc['Hombre']
df_m=df2.loc['Mujer']

df_describe=df_h['Complementos convenio'].describe().to_frame('HOMBRES')
df_describe['MUJERES']=round(df_m['Complementos convenio'].describe().to_frame('MUJERES'),2)

with pd.ExcelWriter(archivo_nombre, engine='openpyxl') as writer2:
  writer2.book = load_workbook(archivo_nombre)
  df_describe.to_excel(writer2, sheet_name="Comp convenio Total")
  writer2.save()


#df_describe.to_excel(writer, engine='openpyxl', sheet_name='Total')

#COMPLEMENTOS NO CONVENIO. TOTAL
df2=df.set_index('Sexo')
df_h=df2.loc['Hombre']
df_m=df2.loc['Mujer']

df_describe=df_h['Complementos no convenio'].describe().to_frame('HOMBRES')
df_describe['MUJERES']=round(df_m['Complementos no convenio'].describe().to_frame('MUJERES'),2)

with pd.ExcelWriter(archivo_nombre, engine='openpyxl') as writer3:
  writer3.book = load_workbook(archivo_nombre)
  df_describe.to_excel(writer3, sheet_name="Comp. no Convenio Total")
  writer3.save()

df_describe

#COMPLEMENTOS NO CONVENIO. MEDIAS POR CATEGORÍA. HOMBRES Y MUJERES
mean_bycat=round(pd.pivot_table(df_h,index=['Categoría'], aggfunc='mean')['Complementos no convenio'].to_frame('Hombres Media'),2)

mean_bycat['Mujeres Media']=round(pd.pivot_table(df_m,index=['Categoría'], aggfunc='mean')['Complementos no convenio'].to_frame('Mujeres Media'),2)
mean_bycat['Dif %']= round((mean_bycat['Hombres Media']/mean_bycat['Mujeres Media'])*100,2)

with pd.ExcelWriter(archivo_nombre, engine='openpyxl') as writer4:
  writer4.book = load_workbook(archivo_nombre)
  mean_bycat.to_excel(writer4, sheet_name="Comp. no Convenio Medias")
  writer4.save()

mean_bycat

#DIFERENCIA ENTRE: COMPLEMENTOS NO CONVENIO - MEDIA (DE LA CATEGORIA) COMPLEMENTOS NO CONVENIO
#cuando no filtramos y dejamos a aquellos trabajadores que no tienen complementos no convenio, las medias nos salen 0. 
filtrados=pd.read_excel('df.xlsx')
filtrados1=round(pd.pivot_table(filtrados, index=['Categoría'],aggfunc='mean')['Diferencia_SC_CNCProm'].to_frame('Brecha complementos no convenio, Promedio'),1)
#filtrados1['Diferencia std']=round(pd.pivot_table(filtrados, index=['Categoría'],aggfunc='std')['Diferencia_SC_CNCProm'].to_frame('Diferencia std'),2)

with pd.ExcelWriter(archivo_nombre, engine='openpyxl') as writer5:
  writer5.book = load_workbook(archivo_nombre)
  filtrados1.to_excel(writer5, sheet_name="Comp. no Convenio Media Diferencias")
  writer5.save()

filtrados1

filtrados.head()

#COMPLEMENTOS NO CONVENIO. DESVIACIONES POR CATEGORÍA. HOMBRES Y MUJERES
std_bycat=round(pd.pivot_table(df_h,index=['Categoría'], aggfunc='std')['Complementos no convenio'].to_frame('Hombres Std'),2)
std_bycat['Mujeres std']=round(pd.pivot_table(df_m,index=['Categoría'], aggfunc='std')['Complementos no convenio'].to_frame('Mujeres Std'),2)

with pd.ExcelWriter(archivo_nombre, engine='openpyxl') as writer6:
  writer6.book = load_workbook(archivo_nombre)
  std_bycat.to_excel(writer6, sheet_name="Comp. no Convenio Desviaciones")
  writer6.save()

std_bycat

df_pesos=df.set_index('Sexo')
df_pesos=df_pesos[['Complementos convenio', 'Complementos no convenio', 'SALARIO BASE', 'Tot. Devengado']]
#filtramos por aquellas personas que tienen complementos de no convenio. Dejamos fuera a aquellas que no tienen estos complementos.
df_pesosconv=df_pesos[df_pesos['Complementos no convenio']!=0]
df_pesos['% Complementos no convenio/Tot. Devengado']=(df_pesos['Complementos no convenio']/df_pesos['Tot. Devengado'])*100
df_pesos['% Complementos convenio/Tot. Devengado']=(df_pesos['Complementos convenio']/df_pesos['Tot. Devengado'])*100
comp_noconvenio_pesos_h=df_pesos[df_pesos['Complementos no convenio']!=0].loc['Hombre']['% Complementos no convenio/Tot. Devengado'].mean(axis=0) 
comp_convenio_pesos_h=df_pesos.loc['Hombre']['% Complementos convenio/Tot. Devengado'].mean(axis=0)
comp_noconvenio_pesos_m=df_pesos[df_pesos['Complementos no convenio']!=0].loc['Mujer']['% Complementos no convenio/Tot. Devengado'].mean(axis=0) 
comp_convenio_pesos_m=df_pesos.loc['Mujer']['% Complementos convenio/Tot. Devengado'].mean(axis=0)

print('Media de complementos no convenio €, HOMBRES', round(df_pesos.loc['Hombre']['Complementos no convenio'].mean(),2))
print('Media de complementos no convenio €, MUJERES', round(df_pesos.loc['Mujer']['Complementos no convenio'].mean(),2), '\n')
print('Media de complementos convenio €, HOMBRES', round(df_pesos.loc['Hombre']['Complementos convenio'].mean(),2))
print('Media de complementos convenio €, MUJERES', round(df_pesos.loc['Mujer']['Complementos convenio'].mean(),2))
print('-------------------------------------------------------------------------')
print('Media de la contribución de los complementos de no convenio sobre el total devengado, HOMBRES:', comp_noconvenio_pesos_h, "%")
print('Media de la contribución de los complementos de no convenio sobre el total devengado, MUJERES:', comp_noconvenio_pesos_m, "%")
print('-------------------------------------------------------------------------')
print('Media de la contribución de los complementos de convenio sobre el total devengado, HOMBRES:', comp_convenio_pesos_h, "%")
print('Media de la contribución de los complementos de convenio sobre el total devengado, MUJERES:', comp_convenio_pesos_m, "%")

#-------------------------------------------------------------------------------------------------------------------------------------------
#GRAFICOS

#se puede ver mayores valores extremos en hombres, mientras que en mujeres hay menos valores extremos pero los sueldos se concentran más en los estratos bajos.
plt.figure(figsize=(15,10))
plt.title('Distribución del total devengado, por género')
sns.violinplot(x='Sexo', y='Tot. Devengado', data=df,orient='v')
#plt.savefig('Violinplot.png')

#variabilidad del total devengado, sin agrupar por categorías
plt.figure(figsize=(15,10))
fig=px.box(df, x='Sexo', y='Tot. Devengado',points='all',hover_data=['Cód. trabajador'],title='Total devengado por género, totales')
fig.show()
#plt.savefig('boxplot.png')

sns.distplot(df_men['Ant. Años'])
plt.title('Años antigüedad, Hombres')

sns.distplot(df_women['Ant. Años'], color='orange')
plt.title('Años antigüedad, Mujeres')

plt.figure(figsize=(15,10))
fig=px.box(df_women, x='Categoría', y='Tot. Devengado',title='Mujeres total devengado, por categorías')
fig.show()
#plt.savefig('boxplot_m.png')

fig=px.box(df_men, x='Categoría', y='Tot. Devengado',title='Hombres total devengado, por categorías')
fig.show()
#plt.savefig('boxplot_h.png')

df_men.groupby(by=['DEPARTAMENTO', 'Cód. trabajador']).sum()

#Representación del total devengado por categorías y sexo. % del total por género
fig = make_subplots(rows=1, cols=2)
fig = px.pie(df_men, values='Tot. Devengado' ,names='Categoría',title='Hombres Total devengado por categoría')
fig2 = px.pie(df_women, values='Tot. Devengado' ,names='Categoría',title='Mujeres Total devengado por categoría')
fig.show()
fig2.show()
#plt.savefig('pieplots.png')

#por categorías, comparación de media Total Devengado entre homvre y mujer
by_cat_sex=df.groupby(by=['Sexo','Categoría']).mean()
by_cat=df.groupby(by=['Categoría','Sexo']).mean()
by_cat_men=by_cat_sex['Tot. Devengado'].loc['Hombre']
by_cat_women=by_cat_sex['Tot. Devengado'].loc['Mujer']

plt.figure(figsize=(20,10))
by_cat_men.plot(kind='bar',label='men')
by_cat_women.plot(kind='bar',color='orange',label='women')
plt.legend()
plt.title('Total devengado por categorías y sexo. PROMEDIO')
#plt.savefig('mean_bycat.png')

#por categorías, comparación principales estadísticos entre hombre y mujer
by_catmen=by_cat['Tot. Devengado'].iloc[by_cat.index.get_level_values('Sexo') == 'Hombre']
by_catwomen=by_cat['Tot. Devengado'].iloc[by_cat.index.get_level_values('Sexo') == 'Mujer']
by_catwomen_describe=by_catwomen.describe(include='all')
by_catmen_describe=by_catmen.describe(include='all')

plt.figure(figsize=(20,10))
by_catmen_describe.plot(kind='bar',label='men')
by_catwomen_describe.plot(kind='bar',color='orange',label='women')
plt.title('Descriptives')
plt.legend()
#plt.savefig('statistics.png')

descriptive=by_catmen_describe.to_frame('Hombres')
descriptive['Mujeres']=by_catwomen_describe
descriptive

from sklearn.model_selection import train_test_split
from sklearn.linear_model import LinearRegression
import numpy as np
from sklearn.metrics import r2_score,mean_squared_error

#df_women or df_men
def regresion(df):
  X = df['Ant. Años']
  y= df['Tot. Devengado']

  X_train,X_test,y_train,y_test = train_test_split(X,y,train_size=0.7,random_state=100)
  X_train = X_train[:,np.newaxis]
  X_test = X_test[:,np.newaxis]

  lr = LinearRegression()
  lr.fit(X_train,y_train)
  y_pred = lr.predict(X_test)

  plt.figure(figsize=(12,6))
  plt.scatter(y_test,y_pred,color='r',linestyle='-')
  plt.show()

  mse = mean_squared_error(y_test,y_pred)
  rsq = r2_score(y_test,y_pred)
  print('mean squared error :',mse)
  print('r square :',rsq)
  print('Intercept of the model:',np.array(lr.intercept_))
  print('Coefficient of the line:',np.array(lr.coef_))

regresion(df_men)

#las mujeres tienen salario más correlacionado con sus años de antigüedad. La antigüedad explica más el salario que para los hombres. En hombres parece que existen otros factores más determinantes.
regresion(df_women)